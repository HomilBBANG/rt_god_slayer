"""
balance.xlsx 초기 파일 생성 스크립트.
이미 파일이 있으면 덮어쓰지 않습니다.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "balance.xlsx")

# ── 색상 ──────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F3864")
CAT_FILL   = PatternFill("solid", fgColor="2E75B6")
KEY_FILL   = PatternFill("solid", fgColor="EBF3FB")
ALT_FILL   = PatternFill("solid", fgColor="F5F9FE")
WHT        = "FFFFFF"
BLK        = "1A1A1A"
thin       = Side(style="thin", color="C0C0C0")
BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_font(bold=True): return Font(name="맑은 고딕", bold=bold, color=WHT, size=10)
def def_font(bold=False): return Font(name="맑은 고딕", bold=bold, color=BLK, size=10)
def center(): return Alignment(horizontal="center", vertical="center")
def left():   return Alignment(horizontal="left",   vertical="center")

def set_col(ws, col, width): ws.column_dimensions[get_column_letter(col)].width = width
def set_row(ws, row, h=18): ws.row_dimensions[row].height = h

def header(ws, texts, row=1):
    for c, t in enumerate(texts, 1):
        cell = ws.cell(row=row, column=c, value=t)
        cell.font      = hdr_font()
        cell.fill      = HDR_FILL
        cell.alignment = center()
        cell.border    = BORDER
    set_row(ws, row, 22)

def write_rows(ws, data, start=2):
    """data: list of (category, key, value, desc, unit)"""
    last_cat = None
    r = start
    for cat, key, val, desc, unit in data:
        # 카테고리 구분행
        if cat != last_cat:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            cell = ws.cell(row=r, column=1, value=f"▶ {cat}")
            cell.font      = Font(name="맑은 고딕", bold=True, color=WHT, size=10)
            cell.fill      = CAT_FILL
            cell.alignment = left()
            cell.border    = BORDER
            set_row(ws, r, 20)
            r += 1
            last_cat = cat

        fill = KEY_FILL if (r - start) % 2 == 0 else ALT_FILL
        for c, v in enumerate([key, val, desc, unit], 2):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill      = fill
            cell.border    = BORDER
            cell.font      = def_font(bold=(c == 2))
            cell.alignment = center() if c in (3, 5) else left()
        set_row(ws, r)
        r += 1
    return r

# ══════════════════════════════════════════════════════════
# 시트 1: Player
# ══════════════════════════════════════════════════════════
PLAYER_DATA = [
    # (카테고리, key, 초기값, 설명, 단위)
    ("HP / SP / MP",  "max_hp",           100,   "최대 HP",                  ""),
    ("HP / SP / MP",  "max_sp",           100,   "최대 SP (스태미나)",         ""),
    ("HP / SP / MP",  "max_mp",            80,   "최대 MP (마나)",             ""),
    ("HP / SP / MP",  "sp_regen",          22,   "SP 초당 자연 회복량",        "/s"),
    ("HP / SP / MP",  "mp_regen",           5,   "MP 초당 자연 회복량",        "/s"),

    ("이동",          "move_speed",        220,   "기본 이동 속도",             "px/s"),
    ("이동",          "crouch_speed_mult",  0.5,  "숙이기 이동속도 배율",        "×"),
    ("이동",          "gravity",           900,   "중력 가속도",               "px/s²"),

    ("점프",          "jump_velocity",    -480,   "점프 초기 속도 (음수=위)",   "px/s"),

    ("대시",          "dash_speed",        650,   "대시 이동 속도",             "px/s"),
    ("대시",          "dash_duration",     0.18,  "대시 지속 시간",             "s"),
    ("대시",          "dash_cooldown",     0.7,   "대시 쿨다운",               "s"),
    ("대시",          "dash_sp_cost",       20,   "대시 SP 소모량",             "SP"),

    ("전투 공통",     "base_atk",           20,   "기본 공격력",               ""),
    ("전투 공통",     "defense",             0,   "방어력 (데미지 감소)",        ""),
    ("전투 공통",     "crit_chance",       0.08,  "크리티컬 확률",              "%"),
    ("전투 공통",     "crit_mult",          1.8,  "크리티컬 배율",              "×"),
    ("전투 공통",     "hit_iframes",        0.5,  "피격 무적 시간",             "s"),
    ("전투 공통",     "knockback_force",   320,   "공격 시 넉백 힘",            "px/s"),

    ("Z 근거리 공격", "melee_combo1_mult",  1.0,  "1타 배율",                  "×"),
    ("Z 근거리 공격", "melee_combo2_mult",  1.3,  "2타 배율",                  "×"),
    ("Z 근거리 공격", "melee_combo3_mult",  1.6,  "3타 배율",                  "×"),
    ("Z 근거리 공격", "melee_range_x",      55,   "근거리 공격 가로 범위",       "px"),
    ("Z 근거리 공격", "melee_range_y",      28,   "근거리 공격 세로 범위",       "px"),
    ("Z 근거리 공격", "melee_atk_cd",      0.28,  "근거리 공격 쿨다운",          "s"),
    ("Z 근거리 공격", "combo_reset_time",   0.6,  "콤보 초기화 시간",            "s"),

    ("X 특수 공격",   "special_atk_mult",   1.7,  "특수 공격 배율",             "×"),
    ("X 특수 공격",   "special_range_x",    85,   "특수 공격 가로 범위",         "px"),
    ("X 특수 공격",   "special_range_y",    36,   "특수 공격 세로 범위",         "px"),
    ("X 특수 공격",   "special_atk_cd",    0.85,  "특수 공격 쿨다운",            "s"),
    ("X 특수 공격",   "special_sp_cost",    25,   "특수 공격 SP 소모",           "SP"),

    ("R 마법",        "magic_atk_mult",     1.2,  "마법 공격 배율",             "×"),
    ("R 마법",        "magic_proj_speed",  480,   "마법 투사체 속도",            "px/s"),
    ("R 마법",        "magic_cd",          0.65,  "마법 쿨다운",               "s"),
    ("R 마법",        "magic_mp_cost",      18,   "마법 MP 소모",              "MP"),
]

# ══════════════════════════════════════════════════════════
# 시트 2: BasicEnemy
# ══════════════════════════════════════════════════════════
ENEMY_DATA = [
    ("기본 스탯", "max_hp",       60,   "최대 HP",              ""),
    ("기본 스탯", "atk",          12,   "공격력",               ""),
    ("기본 스탯", "move_spd",     75,   "이동 속도",             "px/s"),
    ("기본 스탯", "gravity",     900,   "중력 가속도",           "px/s²"),

    ("전투",      "atk_range",    40,   "공격 판정 범위",         "px"),
    ("전투",      "atk_cd",      1.2,   "공격 쿨다운",            "s"),
    ("전투",      "knockback_x", 200,   "공격 시 넉백 X",         "px/s"),
    ("전투",      "knockback_y", -80,   "공격 시 넉백 Y (음수=위)","px/s"),

    ("보상",      "exp_reward",   15,   "처치 시 EXP",            ""),
    ("gold_reward", "gold_reward", 8,   "처치 시 골드",           "G"),
]

# ══════════════════════════════════════════════════════════
# 시트 3: Platforms (장애물/발판)
# ══════════════════════════════════════════════════════════
PLATFORM_DATA = [
    # (이름, pos_x, pos_y, width, height, 비고)
    ("Floor",     640,  680, 2000,  40, "바닥"),
    ("Platform1", 350,  530,  280,  20, "중간 발판 1"),
    ("Platform2", 750,  430,  200,  20, "중간 발판 2"),
    ("Ladder1",   490,  480,   30, 120, "사다리 (Area2D)"),
    ("Enemy1",    500,  640,    0,   0, "적 스폰 위치"),
    ("Enemy2",    750,  640,    0,   0, "적 스폰 위치"),
    ("Enemy3",    900,  640,    0,   0, "적 스폰 위치"),
    ("Player",    200,  600,    0,   0, "플레이어 시작 위치"),
]

# ══════════════════════════════════════════════════════════
# 워크북 생성
# ══════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ── 시트 1: Player ────────────────────────────────────────
ws = wb.active
ws.title = "Player"
ws.freeze_panes = "B2"
header(ws, ["카테고리", "키 (key)", "값 (value)", "설명", "단위"])
set_col(ws, 1, 14); set_col(ws, 2, 24); set_col(ws, 3, 10)
set_col(ws, 4, 28); set_col(ws, 5, 8)
ws.cell(row=1, column=1).value = "카테고리"
write_rows(ws, PLAYER_DATA)

# ── 시트 2: BasicEnemy ───────────────────────────────────
ws2 = wb.create_sheet("BasicEnemy")
ws2.freeze_panes = "B2"
header(ws2, ["카테고리", "키 (key)", "값 (value)", "설명", "단위"])
set_col(ws2, 1, 14); set_col(ws2, 2, 24); set_col(ws2, 3, 10)
set_col(ws2, 4, 28); set_col(ws2, 5, 8)
write_rows(ws2, ENEMY_DATA)

# ── 시트 3: Platforms ────────────────────────────────────
ws3 = wb.create_sheet("Platforms")
ws3.freeze_panes = "A2"
plat_hdr = ["이름 (name)", "pos_x", "pos_y", "width", "height", "비고"]
for c, t in enumerate(plat_hdr, 1):
    cell = ws3.cell(row=1, column=c, value=t)
    cell.font = hdr_font(); cell.fill = HDR_FILL
    cell.alignment = center(); cell.border = BORDER
set_row(ws3, 1, 22)
set_col(ws3, 1, 16); set_col(ws3, 2, 10); set_col(ws3, 3, 10)
set_col(ws3, 4, 10); set_col(ws3, 5, 10); set_col(ws3, 6, 20)

for r, row in enumerate(PLATFORM_DATA, 2):
    fill = KEY_FILL if r % 2 == 0 else ALT_FILL
    for c, v in enumerate(row, 1):
        cell = ws3.cell(row=r, column=c, value=v)
        cell.fill = fill; cell.border = BORDER
        cell.font = def_font(bold=(c == 1))
        cell.alignment = center() if c != 1 else left()
    set_row(ws3, r)

wb.save(OUT)
print(f"✅ 생성 완료: {OUT}")
