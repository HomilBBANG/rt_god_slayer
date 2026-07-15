"""
balance.xlsx → data/balance.json 변환 스크립트.

사용법:
  python tools/excel_to_godot.py
  또는 변환.bat 더블클릭
"""
import os, json, sys
import openpyxl

ROOT  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLS   = os.path.join(ROOT, "tools", "balance.xlsx")
OUT   = os.path.join(ROOT, "data",  "balance.json")

def parse_player(ws):
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key, val = row[1], row[2]  # col B=key, col C=value
        if key and val is not None:
            data[str(key)] = val
    return data

def parse_platforms(ws):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, px, py, w, h, note = row
        if name:
            rows.append({
                "name":   str(name),
                "pos_x":  float(px  or 0),
                "pos_y":  float(py  or 0),
                "width":  float(w   or 0),
                "height": float(h   or 0),
                "note":   str(note or ""),
            })
    return rows

def main():
    if not os.path.exists(XLS):
        print(f"❌ 파일 없음: {XLS}")
        print("   먼저 python tools/create_balance.py 를 실행하세요.")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLS, data_only=True)
    config = {}

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        if sheet == "Platforms":
            config["Platforms"] = parse_platforms(ws)
        else:
            config[sheet] = parse_player(ws)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

    print(f"✅ 변환 완료: {OUT}")
    print(f"   시트: {', '.join(wb.sheetnames)}")
    for sheet, val in config.items():
        count = len(val) if isinstance(val, (dict, list)) else "?"
        print(f"   [{sheet}] {count}개 항목")

if __name__ == "__main__":
    main()
